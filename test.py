from turtle import color
import fitz
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from datetime import datetime
import shutil
from copy import copy
import sys, os
import json

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS   # PyInstaller temp dir
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# ================ to apply icon ===================
def apply_icon(window):
    try:
        window.iconbitmap(resource_path("app-icon.ico"))
    except:
        pass

# ================ handle data types ===================
def to_number(value):
    value = value.strip()
    try:
        return int(value) if value.isdigit() else float(value)
    except ValueError:
        return value



# ================= FILES =================
PDF_IN = None
PDF_OUT = None
doc = None
num_pages = 0
current_page_index = 0
current_page = None  # set after a PDF is opened via open_pdf()
TEMPLATE_XLSX = resource_path("FORMAT.xlsx")

# ================= STATE =================
zoom = 1.5
balloon_no = 1
balloons = []
last_balloon_cache = {"char": "", "req": "", "neg": "", "pos": "", "equip": ""}
current_project_path = None
project_dirty = False
rendered_img = None
rendered_zoom = None
rendered_page_index = None
page_cache = {}  # cache rendered pages per (page_index, zoom)

pan_start = None
offset_x, offset_y = 0, 0
# Two-point balloon placement state
two_point_mode = False
pending_start = None

# =====================================================
# RENDER PDF
# =====================================================
def render_pdf():
    global rendered_img, rendered_zoom, rendered_page_index

    cache_key = (current_page_index, zoom)
    cached = page_cache.get(cache_key)
    if cached:
        rendered_img = cached
    else:
        page = doc[current_page_index]
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        rendered_img = ImageTk.PhotoImage(img)
        page_cache[cache_key] = rendered_img
    rendered_zoom = zoom
    rendered_page_index = current_page_index

    canvas.delete("pdf")
    canvas.create_image(offset_x, offset_y, anchor="nw",
                        image=rendered_img, tags="pdf")

def render_overlays():
    canvas.delete("overlay")

    # Pending start marker for two-point mode
    if pending_start and pending_start["page"] == current_page_index:
        sx = pending_start["x"] * zoom + offset_x
        sy = pending_start["y"] * zoom + offset_y
        marker_r = max(3, (balloon_radius_slider.get() * zoom) / 10)
        canvas.create_oval(
            sx - marker_r, sy - marker_r, sx + marker_r, sy + marker_r,
            outline="red",
            fill="red",
            width=1,
            tags="overlay"
        )

    for b in balloons:
        if b["page"] != current_page_index:
            continue

        x = b["x"] * zoom + offset_x
        y = b["y"] * zoom + offset_y
        r = b["r"] * zoom

        # Draw connector if this balloon was placed via two-point mode
        if b.get("start_x") is not None and b.get("start_y") is not None:
            sx = b["start_x"] * zoom + offset_x
            sy = b["start_y"] * zoom + offset_y
            line_width = max(2, r / 10)
            dx = x - sx
            dy = y - sy
            dist = (dx * dx + dy * dy) ** 0.5
            if dist < 1e-6:
                ex, ey = x, y
            else:
                scale = r / dist
                ex = x - dx * scale
                ey = y - dy * scale
            canvas.create_line(
                sx, sy, ex, ey,
                fill="red",
                width=line_width,
                tags="overlay"
            )
            handle_r = max(3, line_width)
            canvas.create_oval(
                sx - handle_r, sy - handle_r, sx + handle_r, sy + handle_r,
                outline="red",
                fill="red",
                width=1,
                tags="overlay"
            )

        if b.get("highlight"):
            outline = "red"
            fill = "SkyBlue"
            width = max(3, r / 6)
        else:
            outline = "red"
            fill = ""
            width = max(2, r / 10)

        canvas.create_oval(
            x - r, y - r, x + r, y + r,
            outline=outline,
            fill=fill,
            width=width,
            tags="overlay"
        )

        canvas.create_text(
            x, y,
            text=str(b["no"]),
            font=("Arial", int(r)),
            fill=outline,
            tags="overlay"
        )


def render(force=False):
    if not doc:
        return

    if force or rendered_zoom != zoom or rendered_page_index != current_page_index:
        render_pdf()

    render_overlays()
    update_balloon_list()




def open_pdf():
    global PDF_IN, doc, num_pages, current_page_index, balloons, balloon_no
    global offset_x, offset_y, page_cache, pending_start, project_dirty, current_project_path

    path = filedialog.askopenfilename(
        title="Select PDF",
        filetypes=[("PDF files", "*.pdf")]
    )

    if not path:
        return

    if doc:
        doc.close()

    PDF_IN = path
    doc = fitz.open(PDF_IN)
    num_pages = len(doc)
    current_page_index = 0
    balloons.clear()
    balloon_no = 1
    offset_x = offset_y = 0
    page_cache.clear()
    pending_start = None
    
    # Fresh start - no project file associated, clean state
    current_project_path = None
    project_dirty = False

    render(force=True)
    update_two_point_ui()



# =====================================================
# POPUP (Requirement / Tolerance)
# =====================================================

def safe_insert(entry, value): #function to ensure values are stored safely in string form only
    entry.insert(0, "" if value is None else str(value))

def requirement_popup(existing=None):
    if not doc:
        messagebox.showwarning("No File", "Open file to work on")
        return

    popup = tk.Toplevel(root)
    popup.title("Edit FAIR Dimension" if existing else "FAIR Dimension Entry")
    apply_icon(popup)
    popup.transient(root)
    popup.grab_set()

    # Let the grid stretch so buttons can anchor to opposite sides
    popup.columnconfigure(0, weight=1)
    popup.columnconfigure(1, weight=1)

    result = {"action": None}

    # Placeholder/cache helpers (only used for new balloons)
    placeholder_state = {}

    def set_placeholder(widget, text):
        """Show grey placeholder text; mark widget as placeholder."""
        placeholder_state[widget] = {
            "placeholder": bool(text),
            "text": text,
            "pending_clear": False,
        }
        widget.delete(0, tk.END)
        if text:
            widget.insert(0, text)
            _set_fg(widget, "gray")
        else:
            _set_fg(widget, "black")

    def _set_fg(widget, color):
        try:
            widget.configure(foreground=color)
        except Exception:
            try:
                widget.configure(fg=color)
            except Exception:
                pass

    def ensure_state(widget):
        return placeholder_state.get(widget)

    def on_focus_in(event):
        st = ensure_state(event.widget)
        if not st:
            return
        # Keep placeholder visible on focus until user tabs to accept or types.
        if st["placeholder"]:
            _set_fg(event.widget, "gray")

    def on_key_press(event):
        st = ensure_state(event.widget)
        if not st:
            return
        # Ignore control keys that shouldn't clear placeholders
        if event.keysym in ("Tab", "Shift_L", "Shift_R", "Control_L", "Control_R", "Alt_L", "Alt_R", "Return"):
            return
        if st["placeholder"] or st["pending_clear"]:
            event.widget.delete(0, tk.END)
            st["placeholder"] = False
            st["pending_clear"] = False
            _set_fg(event.widget, "black")

    def on_tab(event):
        st = ensure_state(event.widget)
        if not st:
            return
        if st["placeholder"]:
            st["placeholder"] = False
            st["pending_clear"] = False
            _set_fg(event.widget, "black")
        return "break"

    def on_focus_out(event):
        st = ensure_state(event.widget)
        if not st:
            return
        if not event.widget.get().strip() and st["text"]:
            set_placeholder(event.widget, st["text"])

    def on_combobox_select(event):
        st = ensure_state(event.widget)
        if not st:
            return
        st["placeholder"] = False
        st["pending_clear"] = False
        _set_fg(event.widget, "black")

    def bind_placeholder(widget, text):
        set_placeholder(widget, text)
        widget.bind("<FocusIn>", on_focus_in)
        widget.bind("<KeyPress>", on_key_press)
        widget.bind("<FocusOut>", on_focus_out)
        widget.bind("<Tab>", on_tab)  # allow tabbing while accepting placeholder
        if isinstance(widget, ttk.Combobox):
            widget.bind("<<ComboboxSelected>>", on_combobox_select)

    def get_value(widget):
        st = ensure_state(widget)
        if st and st["placeholder"]:
            return ""
        return widget.get().strip()


    tk.Label(popup, text="Characteristic Designator").grid(row=0, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="Requirement").grid(row=1, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="- Tol").grid(row=2, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="+ Tol").grid(row=3, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="Equipment Used").grid(row=4, column=0, padx=8, pady=5, sticky="e")

    char = ttk.Combobox(
        popup,
        values = [
            "--- BASIC DIMENSIONS ---",
            "Length",
            "Width",
            "Thickness",
            "Diameter",
            "Radius",
            "Chamfer",
            "Angle",

            "--- SLOT / CUTOUT ---",
            "Slot length",
            "Slot width",

            "--- HOLE / LOCATION ---",
            "Edge to edge",
            "Edge to hole center",
            "Hole center to edge",
            "Hole center to hole center",
            "Edge to slot",
            "Slot to edge",

            "--- BEND ---",
            "Edge to bend",
            "Bend to edge",
            "Bend to bend",
            "Bending radius",

            "--- EXTRUSION ---",
            "Extrusion height",
            "Extrusion diameter",

            "--- FASTENING ---",
            "Tapping",
            "Riveting",
            "PEM",
            "Torque test",

            "--- FINISH ---",
            "Coating thickness",
            "Aesthetic parameters",

            "--- GD&T ---",
            "Flatness",
            "Parallelism",
            "Perpendicularity",
            "Concentricity",

            "--- MISC ---",
            "Dimension",
            "Ref dimension",
            "Note",
        ],
        width=28
    )

    req = tk.Entry(popup, validate="key")
    neg = tk.Entry(popup, validate="key")
    pos = tk.Entry(popup, validate="key")

    equip = ttk.Combobox(
        popup,
        values = [
            "--- BASIC TOOLS ---",
            "Visual",
            "Vernier caliper",
            "Digital vernier caliper",
            "Digital micrometer",

            "--- HEIGHT MEASUREMENT ---",
            "Digital height gauge",
            "Micro height gauge",

            "--- REFERENCE ---",
            "Slip gauge",

            "--- FORM & ANGLE ---",
            "Radius gauge",
            "Feeler gauge",
            "Bevel protractor",

            "--- GO / NO-GO ---",
            "Pin gauge",
            "Thread plug gauge",

            "--- SURFACE ---",
            "DFT meter",

            "--- ASSEMBLY ---",
            "Torque wrench",

            "--- ADVANCED ---",
            "Profile projector",
            "CMM",
        ],
        width=28
    )

    if existing:
        # print(existing)
        char.insert(0, existing["char"])
        req.insert(0, existing["req"])
        neg.insert(0, existing["neg"])
        pos.insert(0, existing["pos"])
        equip.insert(0, existing["equip"])
        _set_fg(char, "black")
        _set_fg(req, "black")
        _set_fg(neg, "black")
        _set_fg(pos, "black")
        _set_fg(equip, "black")
    else:
        bind_placeholder(char, last_balloon_cache.get("char", ""))
        bind_placeholder(req, last_balloon_cache.get("req", ""))
        bind_placeholder(neg, last_balloon_cache.get("neg", ""))
        bind_placeholder(pos, last_balloon_cache.get("pos", ""))
        bind_placeholder(equip, last_balloon_cache.get("equip", ""))

    char.grid(row=0, column=1, padx=(0, 16))
    req.grid(row=1, column=1, sticky="w")
    neg.grid(row=2, column=1, sticky="w")
    pos.grid(row=3, column=1, sticky="w")
    equip.grid(row=4, column=1, sticky="w")

    # moves the cursor to the input for ease of use
    char.focus_set()
    popup.after(50, lambda: char.select_range(0, tk.END))

    def save():
        req_val_raw = get_value(req)
        neg_val_raw = get_value(neg)
        pos_val_raw = get_value(pos)
        req_val = to_number(req_val_raw)
        neg_val = to_number(neg_val_raw)
        pos_val = to_number(pos_val_raw)

        if not req_val_raw:
            messagebox.showwarning("Missing", "Requirement is mandatory")
            return
        result.update({
            "action": "save",
            "char": get_value(char),
            "req": req_val,
            "neg": neg_val,
            "pos": pos_val,
            "equip": get_value(equip)
        })
        # Update cache on save for new balloons
        if not existing:
            def keep_or(old, new):
                return new if new else old
            last_balloon_cache.update({
                "char": keep_or(last_balloon_cache.get("char", ""), result["char"]),
                "req": keep_or(last_balloon_cache.get("req", ""), req_val_raw),
                "neg": keep_or(last_balloon_cache.get("neg", ""), neg_val_raw),
                "pos": keep_or(last_balloon_cache.get("pos", ""), pos_val_raw),
                "equip": keep_or(last_balloon_cache.get("equip", ""), result["equip"]),
            })
        # print(f"char: {char.get().strip()}, req: {req_val}, neg: {neg_val}, pos: {pos_val}, equip: {equip.get().strip()} \n")
        popup.destroy()

    def delete():
        if messagebox.askyesno("Delete", "Delete this balloon?"):
            result["action"] = "delete"
            popup.destroy()

    if not existing:
        tk.Label(popup, text="Press TAB to autofill", fg="gray").grid(
            row=5, column=0, columnspan=2, sticky="w", padx=(20, 0)
        )
        tk.Button(popup, text="Save", width=10, command=save).grid(
            row=5, column=1, padx=(0, 20), pady=8, sticky="e"
        )

    if existing:
        tk.Button(popup, text="Save", width=10, command=save).grid(
            row=5, column=0, padx=(20, 0), pady=8, sticky="w"
        )
        tk.Button(popup, text="Delete", width=10, fg="red", command=delete).grid(
            row=5, column=1, padx=(0, 20), pady=8, sticky="e"
        )

    # key binds for ease of use
    char.bind("<Return>", lambda e: req.focus_set())

    req.bind("<Return>", lambda e: neg.focus_set())
    req.bind("<Up>", lambda e: char.focus_set())
    req.bind("<Down>", lambda e: neg.focus_set())

    neg.bind("<Return>", lambda e: pos.focus_set())
    neg.bind("<Up>", lambda e: req.focus_set())
    neg.bind("<Down>", lambda e: pos.focus_set())

    pos.bind("<Return>", lambda e: equip.focus_set())
    pos.bind("<Up>", lambda e: neg.focus_set())
    pos.bind("<Down>", lambda e: equip.focus_set())

    equip.bind("<Return>", lambda e: save())
    equip.bind("<Up>", lambda e: pos.focus_set())

    popup.bind("<Escape>", lambda e: popup.destroy())

    popup.wait_window()
    return result



# =====================================================
# ADD balloon (IMMEDIATE)
# =====================================================
def clear_pending_start():
    global pending_start
    pending_start = None
    render_overlays()
    update_two_point_ui()


def toggle_two_point_mode():
    global two_point_mode
    two_point_mode = not two_point_mode
    clear_pending_start()
    update_two_point_ui()


def add_balloon(event):
    if not doc:
        messagebox.showwarning("No File", "Open file to work on")
        return

    global balloon_no, pending_start

    pdf_x = (event.x - offset_x) / zoom
    pdf_y = (event.y - offset_y) / zoom

    # Two-click flow
    if two_point_mode:
        if pending_start and pending_start["page"] != current_page_index:
            messagebox.showinfo("Two-Point Mode", "Start point was on another page. Starting over on this page.")
            pending_start = None

        if not pending_start:
            pending_start = {"page": current_page_index, "x": pdf_x, "y": pdf_y}
            render_overlays()
            update_two_point_ui()
            return
        else:
            start = pending_start
            pending_start = None

    balloons.append({
        "page": current_page_index,
        "no": balloon_no,
        "x": pdf_x,
        "y": pdf_y,
        "r": balloon_radius_slider.get(),
        "char": "",
        "req": "",
        "neg": "",
        "pos": "",
        "equip": "",
        "highlight": False,
        "start_x": start["x"] if two_point_mode else None,
        "start_y": start["y"] if two_point_mode else None,
    })


    render()

    data = requirement_popup()

    # requirement_popup returns {"action": "save"|"delete"|None, ...}
    if data.get("action") == "save":
        balloons[-1]["char"] = data["char"]
        balloons[-1]["req"]  = data["req"]
        balloons[-1]["neg"]  = data["neg"]
        balloons[-1]["pos"]  = data["pos"]
        balloons[-1]["equip"]  = data["equip"]
        balloon_no += 1
        
        # Mark project as dirty
        global project_dirty
        project_dirty = True
    else:
        # If the dialog was closed or cancelled, discard the pending balloon
        balloons.pop()
        if two_point_mode:
            clear_pending_start()

    render()
    update_two_point_ui()


# =====================================================
# DELETE balloon 
# =====================================================
def delete_balloon(balloon):
    balloons.remove(balloon)

    # renumber globally
    for i, b in enumerate(balloons, start=1):
        b["no"] = i

    global balloon_no, project_dirty
    balloon_no = len(balloons) + 1
    project_dirty = True

    update_balloon_list()
    render(force=True)



def highlight_balloon(balloon, duration=2000):

    balloon["highlight"] = True
    render_overlays()

    def clear():
        balloon["highlight"] = False
        render_overlays()

    root.after(duration, clear)

# =====================================================
# EDIT balloon (from list)
# =====================================================
def on_balloon_edit(balloon):

    highlight_balloon(balloon)

    result = requirement_popup(existing=balloon)

    if result["action"] == "save":
        balloon["char"] = result["char"]
        balloon["req"]  = result["req"]
        balloon["neg"]  = result["neg"]
        balloon["pos"]  = result["pos"]
        balloon["equip"]  = result["equip"]
        
        # Mark project as dirty
        global project_dirty
        project_dirty = True

    elif result["action"] == "delete":
        delete_balloon(balloon)
        return

    update_balloon_list()
    render(force=True)

def on_balloon_edit_mouse(event):
    lb = event.widget
    idx = lb.nearest(event.y)

    if idx < 2:
        return

    lb.selection_clear(0, tk.END)
    lb.selection_set(idx)
    lb.activate(idx)

    page_balloons = [b for b in balloons if b["page"] == current_page_index]
    balloon_idx = idx - 2

    if balloon_idx >= len(page_balloons):
        return

    on_balloon_edit(page_balloons[balloon_idx])


def on_balloon_edit_key(event):
    lb = event.widget
    sel = lb.curselection()

    if not sel:
        return

    idx = sel[0]
    if idx < 2:
        return

    page_balloons = [b for b in balloons if b["page"] == current_page_index]
    balloon_idx = idx - 2

    if balloon_idx >= len(page_balloons):
        return

    on_balloon_edit(page_balloons[balloon_idx])


def on_balloon_delete_key(event):
    lb = event.widget
    sel = lb.curselection()

    if not sel:
        return

    idx = sel[0]
    if idx < 2:
        return

    page_balloons = [b for b in balloons if b["page"] == current_page_index]
    balloon_idx = idx - 2

    if balloon_idx >= len(page_balloons):
        return

    balloon = page_balloons[balloon_idx]

    if messagebox.askyesno("Delete", "Delete this balloon?"):
        delete_balloon(balloon)


# =====================================================
# ZOOM / PAN
# =====================================================
zoom_job = None

def apply_zoom(factor, center=None):
    """Shared zoom logic for mouse wheel and keybinds."""
    global zoom, offset_x, offset_y, zoom_job, page_cache

    new_zoom = zoom * factor
    if new_zoom > 10.0 or new_zoom < 0.5:
        return

    if center is None:
        mx = canvas.winfo_width() / 2
        my = canvas.winfo_height() / 2
    else:
        mx, my = center

    offset_x = mx - factor * (mx - offset_x)
    offset_y = my - factor * (my - offset_y)
    if new_zoom != zoom:
        page_cache.clear()
    zoom = new_zoom

    update_preview(balloon_radius_slider.get())

    if zoom_job:
        root.after_cancel(zoom_job)

    zoom_job = root.after(120, lambda: render(force=True))
    update_zoom_ui(zoom)

def zoom_canvas(event):
    factor = 1.25 if event.delta > 0 else 1 / 1.25
    apply_zoom(factor, center=(event.x, event.y))

def zoom_in_key(event=None):
    apply_zoom(1.25)

def zoom_out_key(event=None):
    apply_zoom(1 / 1.25)

# =====================================================
# balloon size via keyboard
# =====================================================
def _set_balloon_radius(delta):
    val = balloon_radius_slider.get() + delta
    val = max(balloon_radius_slider.cget("from"), min(balloon_radius_slider.cget("to"), val))
    balloon_radius_slider.set(val)
    update_preview(val)

def radius_increase(event=None):
    _set_balloon_radius(1)

def radius_decrease(event=None):
    _set_balloon_radius(-1)


def start_pan(event):
    global pan_start
    pan_start = (event.x, event.y)

def do_pan(event):
    global offset_x, offset_y, pan_start
    if pan_start:
        dx = event.x - pan_start[0]
        dy = event.y - pan_start[1]
        offset_x += dx
        offset_y += dy
        pan_start = (event.x, event.y)

        canvas.move("pdf", dx, dy)
        canvas.move("overlay", dx, dy)


def end_pan(event):
    global pan_start
    pan_start = None

# =====================================================
# PAGE NAV / UNDO
# =====================================================
def next_page():
    global current_page_index, offset_x, offset_y
    if current_page_index < num_pages - 1:
        current_page_index += 1
        offset_x = offset_y = 0
        clear_pending_start()
        render()

def prev_page():
    global current_page_index, offset_x, offset_y
    if current_page_index > 0:
        current_page_index -= 1
        offset_x = offset_y = 0
        clear_pending_start()
        render()

def undo():
    global balloon_no, project_dirty
    for i in reversed(range(len(balloons))):
        if balloons[i]["page"] == current_page_index:
            balloons.pop(i)
            balloon_no -= 1
            project_dirty = True
            break
    render()


# =====================================================
# SHORTCUTS MENU
# =====================================================
def show_shortcuts():
    win = tk.Toplevel(root)
    win.title("Keyboard Shortcuts")
    win.transient(root)
    win.grab_set()
    apply_icon(win)

    frame = tk.Frame(win, padx=10, pady=10)
    frame.pack(fill="both", expand=True)

    shortcuts = [
        ("Ctrl + O", "Open PDF"),
        ("Ctrl + P", "Open Project (.fairy)"),
        ("Ctrl + Shift + P", "Save Project (.fairy)"),
        ("Ctrl + S", "Save PDF"),
        ("Ctrl + Shift + S", "Save Report"),
        ("Escape", "Exit any Popup"),
        ("Ctrl + Q", "Exit Application"),
        ("Ctrl + T", "Toggle balloon Mode"),
        ("Ctrl + Z", "Undo balloon"),
        ("Right-Click x2", "Two-point mode: start then end point"),
        ("Enter", "Edit Selected balloon"),
        ("Delete", "Delete Selected balloon"),
        ("Shift + ↑ / ↓", "Change balloon Size"),
        ("← / →", "Prev / Next Page"),
        ("↑ / ↓", "Prev / Next List Item"),
        ("Ctrl + + / -", "Zoom In / Out"),
        ("Ctrl + /", "Show Shortcuts"),
    ]

    for i, (key, action) in enumerate(shortcuts):
        tk.Label(frame, text=key, font=("Consolas", 11, "bold")).grid(row=i, column=0, sticky="w", padx=(0, 20))
        tk.Label(frame, text=action, font=("Segoe UI", 11)).grid(row=i, column=1, sticky="w")

    win.bind("<Return>", lambda e: win.destroy())
    win.bind("<Escape>", lambda e: win.destroy())

    win.focus_set()




# =====================================================
# LIST VIEW
# =====================================================
def update_balloon_list():
    # Fixed column widths for neat alignment in the list view
    w_no, w_char, w_req, w_tol, w_equip = 3, 28, 8, 6, 22
    header = (
        f"{'No':<{w_no}} | "
        f"{'Char':<{w_char}} | "
        f"{'Req':<{w_req}} | "
        f"{'-Tol':<{w_tol}} | "
        f"{'+Tol':<{w_tol}} | "
        f"{'Equip':<{w_equip}}"
    )
    sep = "-" * len(header)

    balloon_listbox.delete(0, tk.END)
    balloon_listbox.insert(tk.END, header)
    balloon_listbox.insert(tk.END, sep)

    for b in balloons:
        if b["page"] == current_page_index:
            balloon_listbox.insert(
                tk.END,
                f"{str(b['no']):<{w_no}} | "
                f"{str(b['char']):<{w_char}} | "
                f"{str(b['req']):<{w_req}} | "
                f"{str(b['neg']):<{w_tol}} | "
                f"{str(b['pos']):<{w_tol}} | "
                f"{str(b['equip']):<{w_equip}}"
            )

# =====================================================
# SAVE balloonD PDF
# =====================================================
def save_pdf():
    if not doc:
        messagebox.showwarning("No File", "No file to save PDF")
        return

    if not balloons:
        messagebox.showwarning("No data", "No balloons to export")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    default_name = f"FAIR_QA_drawing_{timestamp}.pdf"

    PDF_OUT = filedialog.asksaveasfilename(
        initialfile=default_name,
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        title="Save QA Report As"
    )

    if not PDF_OUT:
        return

    out = fitz.open()
    for i in range(num_pages):
        p = out.new_page(width=doc[i].rect.width, height=doc[i].rect.height)
        p.show_pdf_page(p.rect, doc, i)
        for b in balloons:
            if b["page"] == i:
                x, y, r = b["x"], b["y"], b["r"]
                # Draw connector if present (project to balloon edge)
                if b.get("start_x") is not None and b.get("start_y") is not None:
                    sx, sy = b["start_x"], b["start_y"]
                    dx = x - sx
                    dy = y - sy
                    dist = (dx * dx + dy * dy) ** 0.5
                    if dist < 1e-6:
                        ex, ey = x, y
                    else:
                        scale = r / dist
                        ex = x - dx * scale
                        ey = y - dy * scale
                    p.draw_line(fitz.Point(sx, sy), fitz.Point(ex, ey), color=(1, 0, 0), width=r/10)
                    handle_r = r/10
                    p.draw_oval(
                        fitz.Rect(sx - handle_r, sy - handle_r, sx + handle_r, sy + handle_r),
                        color=(1, 0, 0),
                        fill=(1, 0, 0),
                        width=1
                    )
                p.draw_oval(fitz.Rect(x-r, y-r, x+r, y+r), color=(1,0,0), width=r/10)
                text = str(b["no"])
                font_size = r
                try:
                    text_width = fitz.get_text_length(text, fontsize=font_size)
                except Exception:
                    text_width = font_size * 0.6 * len(text)  # fallback estimate
                # center horizontally; adjust baseline to visually center vertically
                tx = x - text_width / 2
                ty = y + font_size * 0.35
                p.insert_text(fitz.Point(tx, ty), text, fontsize=font_size, color=(1,0,0))
    out.save(PDF_OUT)
    out.close()
    messagebox.showinfo("Saved", f"Ballooned drawing saved as {PDF_OUT}")
    try:
        if messagebox.askyesno("Open file", "Open the saved PDF now?"):
            os.startfile(PDF_OUT)
    except Exception:
        pass


# =====================================================
# SAVE REPORT (UNCHANGED)
# =====================================================
def save_report():
    if not doc:
        messagebox.showwarning("No File", "No file to save Report")
        return

    if not balloons:
        messagebox.showwarning("No data", "No balloons to export")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    default_name = f"FAIR_Report_{timestamp}.xlsx"

    report_file = filedialog.asksaveasfilename(
        title="Save FAIR Report",
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not report_file:
        return

    shutil.copy(TEMPLATE_XLSX, report_file)
    wb = load_workbook(report_file)
    ws = wb.active

    # The template has 15 data rows (8-22). Row 23 participates in the merged
    # Legends block (A23:B24, C23:M24), so we must insert before that at row 23.
    START_ROW = 8
    TEMPLATE_ROWS = 15
    LAST_TEMPLATE_ROW = START_ROW + TEMPLATE_ROWS - 1  # 22
    INSERT_AT = LAST_TEMPLATE_ROW + 1                  # 23 (just before merged legends)
    template_row_idx = LAST_TEMPLATE_ROW
    footer_rows = {24: ws.row_dimensions[24].height,
                   25: ws.row_dimensions[25].height,
                   26: ws.row_dimensions[26].height}

    def copy_row_style(src_row_idx, dest_row_idx, clear_values=True):
        # Copy styles across all columns to preserve borders/alignment
        for col_idx in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=src_row_idx, column=col_idx)
            dest_cell = ws.cell(row=dest_row_idx, column=col_idx)
            dest_cell._style = copy(src_cell._style)
            if clear_values:
                dest_cell.value = None
        # Copy row height to keep grid consistent
        ws.row_dimensions[dest_row_idx].height = ws.row_dimensions[src_row_idx].height

    balloon_count = len(balloons)
    extra_rows = max(0, balloon_count - TEMPLATE_ROWS)

    # Insert new rows just above the merged footer block (row 23+)
    if extra_rows > 0:
        ws.insert_rows(INSERT_AT, amount=extra_rows)

        for i in range(extra_rows):
            target_row = INSERT_AT + i
            copy_row_style(template_row_idx, target_row)

        # Re-merge the footer block (Legends/Observations/Sign-off) so it
        # stays intact after the shift, and restore footer row heights.
        footer_merges = [
            "A24:B24", "C24:M24",
            "D25:M25",
            "A26:B26", "C26:D26", "E26:F26", "G26:H26", "I26:J26", "K26:M26",
        ]
        for rng in footer_merges:
            if rng in ws.merged_cells:
                ws.unmerge_cells(rng)
        for rng in footer_merges:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            ws.merge_cells(
                start_row=min_row + extra_rows,
                start_column=min_col,
                end_row=max_row + extra_rows,
                end_column=max_col,
            )
        for src_row, height in footer_rows.items():
            ws.row_dimensions[src_row + extra_rows].height = height

    # Refresh styles on all data rows (including originals) to keep borders/alignments
    style_src_row = START_ROW  # first data row has canonical formatting
    total_rows = balloon_count + 1
    for r in range(START_ROW, START_ROW + total_rows):
        copy_row_style(style_src_row, r, clear_values=False)

    # WRITE DATA (flows naturally)
    row = START_ROW
    for b in balloons:
        ws.cell(row=row, column=1).value = b["page"] + 1
        ws.cell(row=row, column=2).value = b["no"]
        ws.cell(row=row, column=3).value = b['char']
        ws.cell(row=row, column=4).value = b["req"]
        ws.cell(row=row, column=5).value = b["neg"]
        ws.cell(row=row, column=6).value = b["pos"]
        ws.cell(row=row, column=7).value = b["equip"]
        row += 1

    wb.save(report_file)
    messagebox.showinfo("Saved", f"FAIR report created:\n{report_file}")
    try:
        if messagebox.askyesno("Open file", "Open the saved report now?"):
            os.startfile(report_file)
    except Exception:
        pass


# =====================================================
# SAVE PROJECT (.fairy)
# =====================================================
def save_project_to_path(project_file):
    """Save project to a specific path without dialog. Returns True on success."""
    if not doc or not balloons:
        return False
    
    # Build project data structure
    project_data = {
        "version": 1,
        "pdf": {
            "path": PDF_IN,
            "page_count": num_pages
        },
        "balloons": []
    }

    # Deep copy balloons, excluding UI-only fields
    for b in balloons:
        balloon_data = {
            "page": b["page"],
            "no": b["no"],
            "x": b["x"],
            "y": b["y"],
            "r": b["r"],
            "char": b["char"],
            "req": b["req"],
            "neg": b["neg"],
            "pos": b["pos"],
            "equip": b["equip"]
        }
        # Include connector data if present
        if b.get("start_x") is not None:
            balloon_data["start_x"] = b["start_x"]
        if b.get("start_y") is not None:
            balloon_data["start_y"] = b["start_y"]
        
        project_data["balloons"].append(balloon_data)

    try:
        with open(project_file, 'w') as f:
            json.dump(project_data, f, indent=2)
        
        # Track current project for session persistence and mark clean
        global current_project_path, project_dirty
        current_project_path = project_file
        project_dirty = False

        # Save state immediately
        state = {"last_project": project_file}
        save_app_state(state)

        return True
    except Exception:
        return False


def save_project():
    """Save project with dialog (Save As behavior)."""
    if not doc:
        messagebox.showwarning("No File", "No file to save project")
        return False

    if not balloons:
        messagebox.showwarning("No data", "No balloons to save")
        return False

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    default_name = f"FAIR_Project_{timestamp}.fairy"

    project_file = filedialog.asksaveasfilename(
        title="Save Project As",
        initialfile=default_name,
        defaultextension=".fairy",
        filetypes=[("FAIR Project files", "*.fairy"), ("All files", "*.*")]
    )

    if not project_file:
        return False
    
    if save_project_to_path(project_file):
        messagebox.showinfo("Saved", f"Project saved successfully:\n{project_file}")
        return True
    else:
        messagebox.showerror("Save Error", "Failed to save project")
        return False


# =====================================================
# LOAD PROJECT (.fairy)
# =====================================================
def load_project_from_path(project_file, show_success_msg=True, prompt_for_pdf=True):
    """Core project loading logic (no file dialog). Returns True on success, False on failure."""
    try:
        with open(project_file, 'r') as f:
            project_data = json.load(f)
    except json.JSONDecodeError:
        messagebox.showerror("Load Error", "Invalid project file: corrupted or not valid JSON")
        return False
    except Exception as e:
        messagebox.showerror("Load Error", f"Failed to load project:\n{str(e)}")
        return False

    # Validate version
    if project_data.get("version") != 1:
        messagebox.showerror("Unsupported Version", 
                            f"This project was created with version {project_data.get('version')}.\n"
                            "Only version 1 is supported.")
        return False

    # Validate required keys
    if "pdf" not in project_data or "balloons" not in project_data:
        messagebox.showerror("Invalid Project", "Project file is missing required data")
        return False

    pdf_info = project_data["pdf"]
    if "path" not in pdf_info or "page_count" not in pdf_info:
        messagebox.showerror("Invalid Project", "Project file is missing PDF information")
        return False

    # Handle PDF path
    pdf_path = pdf_info["path"]
    if not os.path.exists(pdf_path):
        if not prompt_for_pdf:
            return False
        
        response = messagebox.askokcancel(
            "PDF Not Found",
            f"Original PDF not found:\n{pdf_path}\n\nWould you like to locate it?"
        )
        if not response:
            return False
        
        pdf_path = filedialog.askopenfilename(
            title="Locate Original PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        
        if not pdf_path:
            return False

    # Close existing document if open
    global doc, PDF_IN, num_pages, current_page_index, balloons, balloon_no
    global zoom, offset_x, offset_y, page_cache, pending_start, project_dirty

    if doc:
        doc.close()

    # Open the PDF
    try:
        doc = fitz.open(pdf_path)
        PDF_IN = pdf_path
        num_pages = len(doc)
    except Exception as e:
        messagebox.showerror("PDF Error", f"Failed to open PDF:\n{str(e)}")
        doc = None
        return False

    # Validate page count matches
    if num_pages != pdf_info["page_count"]:
        messagebox.showwarning(
            "Page Count Mismatch",
            f"Warning: PDF has {num_pages} pages but project expected {pdf_info['page_count']}"
        )

    # Load balloons
    balloons.clear()
    skipped_balloons = []

    for balloon_data in project_data["balloons"]:
        # Validate balloon has required fields
        required_fields = ["page", "no", "x", "y", "r", "char", "req", "neg", "pos", "equip"]
        if not all(field in balloon_data for field in required_fields):
            skipped_balloons.append(f"Balloon {balloon_data.get('no', '?')} - missing fields")
            continue

        # Skip balloons referencing invalid pages
        if balloon_data["page"] >= num_pages or balloon_data["page"] < 0:
            skipped_balloons.append(f"Balloon {balloon_data['no']} - invalid page {balloon_data['page']}")
            continue

        # Create balloon with all data
        balloon = {
            "page": balloon_data["page"],
            "no": balloon_data["no"],
            "x": balloon_data["x"],
            "y": balloon_data["y"],
            "r": balloon_data["r"],
            "char": balloon_data["char"],
            "req": balloon_data["req"],
            "neg": balloon_data["neg"],
            "pos": balloon_data["pos"],
            "equip": balloon_data["equip"],
            "highlight": False,  # UI state not saved
            "start_x": balloon_data.get("start_x"),
            "start_y": balloon_data.get("start_y")
        }
        balloons.append(balloon)

    # Recalculate balloon number
    balloon_no = len(balloons) + 1

    # Reset session state
    current_page_index = 0
    zoom = 1.5
    offset_x = offset_y = 0
    page_cache.clear()
    pending_start = None

    # Render the first page
    render(force=True)
    update_two_point_ui()
    
    # Mark project as clean after successful load
    project_dirty = False

    # Show warnings if any balloons were skipped
    if show_success_msg:
        if skipped_balloons:
            messagebox.showwarning(
                "Load Warning",
                f"Project loaded, but {len(skipped_balloons)} balloon(s) were skipped:\n\n" +
                "\n".join(skipped_balloons[:5]) +
                (f"\n... and {len(skipped_balloons) - 5} more" if len(skipped_balloons) > 5 else "")
            )
        else:
            messagebox.showinfo("Loaded", f"Project loaded successfully:\n{len(balloons)} balloon(s) restored")
    
    return True


def load_project():
    """Open project with file dialog."""
    project_file = filedialog.askopenfilename(
        title="Open Project",
        filetypes=[("FAIR Project files", "*.fairy"), ("All files", "*.*")]
    )

    if not project_file:
        return
    
    load_project_from_path(project_file)


# =====================================================
# STATE MANAGEMENT (AppData persistence)
# =====================================================
def get_state_path():
    """Return path to app state file in AppData."""
    app_dir = os.path.join(os.environ.get("APPDATA", ""), "FAIR-y")
    os.makedirs(app_dir, exist_ok=True)
    return os.path.join(app_dir, "state.json")

def load_app_state():
    """Load app state from AppData. Returns empty dict if missing."""
    path = get_state_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r") as f:
            return json.load(f)
    except Exception:
        return {}

def save_app_state(state):
    """Save app state to AppData."""
    try:
        with open(get_state_path(), "w") as f:
            json.dump(state, f, indent=2)
    except Exception:
        pass  # Silent fail - state is non-critical


def auto_restore_last_project():
    """Attempt to restore last project on startup. Fails silently."""
    state = load_app_state()
    last_project = state.get("last_project")
    
    if not last_project:
        return
    
    if not os.path.exists(last_project):
        # Silently skip - file moved/deleted
        return
    
    try:
        # Read project file to check PDF exists
        with open(last_project, 'r') as f:
            project_data = json.load(f)
        
        # Check PDF exists
        pdf_path = project_data.get("pdf", {}).get("path")
        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showinfo(
                "Session Restore",
                f"Could not restore last session:\nPDF file not found.\n\nStarting fresh.",
                parent=root
            )
            return
        
        # Use internal load logic without prompts
        load_project_from_path(last_project, show_success_msg=False, prompt_for_pdf=False)
        
    except Exception:
        # Silent fail - corrupt file or other error
        pass


def on_app_close():
    # Check for unsaved changes
    if not project_dirty:
        root.destroy()
        return
    
    # Ask user what to do with unsaved changes
    choice = messagebox.askyesnocancel(
        "Unsaved Changes",
        "You have unsaved changes to the project.\n\nSave before closing?"
    )
    
    if choice is True:  # Save
        # If project has a path, save silently; otherwise show Save As dialog
        if current_project_path:
            if save_project_to_path(current_project_path):
                root.destroy()
            # If save failed, don't close
        else:
            # No path yet - show Save As dialog
            if save_project():
                root.destroy()
            # If user cancelled save dialog or save failed, don't close
    elif choice is False:  # Don't Save
        root.destroy()
    # else: choice is None (Cancel) - do nothing, stay open


# =====================================================
# UI
# =====================================================
def update_two_point_ui():
    if "two_point_button" not in globals():
        return
    mode_text = "Balloon with Line" if two_point_mode else "Balloon without Line"
    two_point_button.config(
        text=mode_text,
        relief="sunken" if two_point_mode else "raised"
    )
    render_two_point_preview()


root = tk.Tk()
# Start maximized to use full screen
try:
    root.state("zoomed")  # Windows
except Exception:
    root.attributes("-zoomed", True)  # fallback
root.title("FAIR-y")
root.iconbitmap(resource_path("app-icon.ico"))

root.protocol("WM_DELETE_WINDOW", on_app_close)

toolbar = tk.Frame(root)
toolbar.pack(fill="x", padx=5)

#=======================================================
# UI Button Binds
#=======================================================
tk.Button(toolbar, text="Open PDF", command=open_pdf).pack(side="left")
tk.Button(toolbar, text="Open Project", command=load_project).pack(side="left", padx=(0,5))
tk.Button(toolbar, text="Prev Page", command=prev_page).pack(side="left")
tk.Button(toolbar, text="Next Page", command=next_page).pack(side="left", padx=(0,5))
tk.Button(toolbar, text="Undo balloon", command=undo).pack(side="left", padx=(0,5))
tk.Button(toolbar, text="Save PDF", command=save_pdf).pack(side="left")
tk.Button(toolbar, text="Save Report", command=save_report).pack(side="left")
tk.Button(toolbar, text="Save Project", command=save_project).pack(side="left", padx=(0,5))
tk.Button(toolbar, text="Help", command=show_shortcuts).pack(side="left")

def render_two_point_preview():
    if "two_point_preview" not in globals():
        return
    c = two_point_preview
    c.delete("all")
    w = int(c.cget("width"))
    h = int(c.cget("height"))
    padding = 6
    r = 9
    cx = w - padding - r
    cy = h // 2
    outline = "red"
    if two_point_mode:
        sx = padding + 4
        sy = cy
        # project line to circle edge
        dx = cx - sx
        dy = cy - sy
        dist = (dx * dx + dy * dy) ** 0.5 or 1
        ex = cx - dx * (r / dist)
        ey = cy - dy * (r / dist)
        c.create_line(sx, sy, ex, ey, fill=outline, width=2)
        c.create_oval(sx-3, sy-3, sx+3, sy+3, outline=outline, fill=outline, width=1)
    c.create_oval(cx-r, cy-r, cx+r, cy+r, outline=outline, width=2)

#=======================================================
# Keyboard Button Binds
#=======================================================
root.bind("<Control-O>", lambda e: open_pdf())
root.bind("<Control-o>", lambda e: open_pdf())
root.bind("<Control-P>", lambda e: load_project())
root.bind("<Control-p>", lambda e: load_project())
root.bind("<Control-Shift-P>", lambda e: save_project())
root.bind("<Control-Shift-p>", lambda e: save_project())
root.bind("<Control-S>", lambda e: save_pdf())
root.bind("<Control-s>", lambda e: save_pdf())
root.bind("<Control-Shift-S>", lambda e: save_report())
root.bind("<Right>", lambda e: next_page())
root.bind("<Left>", lambda e: prev_page())
root.bind("<Control-Z>", lambda e: undo())
root.bind("<Control-z>", lambda e: undo())
root.bind("<Control-/>", lambda e: show_shortcuts())
root.bind("<Control-plus>", zoom_in_key)
root.bind("<Control-KP_Add>", zoom_in_key)
root.bind("<Control-equal>", zoom_in_key)  # Ctrl+= for + on many keyboards
root.bind("<Control-minus>", zoom_out_key)
root.bind("<Control-KP_Subtract>", zoom_out_key)
root.bind("<Shift-Up>", radius_increase)
root.bind("<Shift-Down>", radius_decrease)
root.bind("<Control-Q>", lambda e: on_app_close())
root.bind("<Control-q>", lambda e: on_app_close())
root.bind("<Control-T>", lambda e: toggle_two_point_mode())
root.bind("<Control-t>", lambda e: toggle_two_point_mode())

#=========================zoom slider===============================
def set_zoom_from_slider(val):
    """Set absolute zoom from the slider (%), updating label and render."""
    global zoom
    try:
        target = float(val) / 100.0
    except Exception:
        return
    target = max(0.5, min(10.0, target))
    if zoom_slider_updating:
        return
    apply_zoom(target / zoom)
    update_zoom_ui(target)

def update_zoom_ui(current_zoom):
    """Keep zoom slider/label in sync after any zoom change (keys/mouse/slider)."""
    pct = int(current_zoom * 100)
    if 'zoom_slider' in globals():
        global zoom_slider_updating
        zoom_slider_updating = True
        try:
            zoom_slider.set(pct)
        finally:
            zoom_slider_updating = False

zoom_slider_updating = False

zoom_slider = tk.Scale(
    toolbar,
    from_=50,
    to=1000,
    orient="horizontal",
    label="Zoom %",
    command=set_zoom_from_slider
)
zoom_slider.set(int(zoom * 100))
zoom_slider.pack(padx=5, side="right")
update_zoom_ui(zoom)

balloon_radius_slider = tk.Scale(toolbar, from_=3, to=25, orient="horizontal", label="Balloon Size")
balloon_radius_slider.set(6)
balloon_radius_slider.pack(side="right")

preview_canvas = tk.Canvas(toolbar, width=50, height=50, bg="#ababab")
preview_canvas.pack(side="right", padx=5)

def update_preview(val):
    preview_canvas.delete("all")
    r = int(val) * zoom
    preview_canvas.create_oval(25-r, 25-r, 25+r, 25+r, outline="red", width=2)

balloon_radius_slider.config(command=update_preview)
update_preview(balloon_radius_slider.get())

#===========================two-point-mode====================================
two_point_button = tk.Button(toolbar, text="Balloon without line", width= 16, command=toggle_two_point_mode)
two_point_button.pack(side="right", padx=(8, 0))
two_point_preview = tk.Canvas(toolbar, width=50, height=50, bg="#ababab", highlightthickness=0)
two_point_preview.pack(side="right", padx=(6, 0))
update_two_point_ui()

# Splitter so the list can be resized by the user
paned = tk.PanedWindow(root, orient="vertical")
paned.pack(fill="both", expand=True)

balloon_listbox = tk.Listbox(
    paned,
    font=("Consolas", 10),
    selectmode="browse",
    exportselection=False
)

balloon_listbox.bind("<Double-Button-1>", on_balloon_edit_mouse)
balloon_listbox.bind("<Return>", on_balloon_edit_key)
balloon_listbox.bind("<Delete>", on_balloon_delete_key)

paned.add(balloon_listbox, minsize=120)

# Give initial focus so arrow keys work without first click
root.after(50, balloon_listbox.focus_set)


canvas = tk.Canvas(paned, bg="gray")
paned.add(canvas, minsize=200)

canvas.bind("<Button-3>", add_balloon)
canvas.bind("<Button-1>", start_pan)
canvas.bind("<B1-Motion>", do_pan)
canvas.bind("<ButtonRelease-1>", end_pan)
canvas.bind("<MouseWheel>", zoom_canvas)

# Auto-restore last session
root.after(100, auto_restore_last_project)

render()
root.mainloop()
if doc:
    doc.close()
