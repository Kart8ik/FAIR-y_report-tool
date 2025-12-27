import fitz
import tkinter as tk
from tkinter import messagebox, filedialog
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from datetime import datetime
import shutil
from copy import copy
import sys, os

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS   # PyInstaller temp dir
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# ================= FILES =================
PDF_IN = None
PDF_OUT = None
doc = None
num_pages = 0
current_page_index = 0
current_page = None  # set after a PDF is opened via open_pdf()
TEMPLATE_XLSX = resource_path("FORMAT.xlsx")

# ================= STATE =================
zoom = 1.5
bubble_no = 1
bubbles = []
rendered_img = None
rendered_zoom = None

pan_start = None
offset_x, offset_y = 0, 0

# =====================================================
# RENDER PDF
# =====================================================
def render_pdf():
    global rendered_img, rendered_zoom

    page = doc[current_page_index]
    mat = fitz.Matrix(zoom, zoom) 
    pix = page.get_pixmap(matrix=mat)

    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    rendered_img = ImageTk.PhotoImage(img)
    rendered_zoom = zoom

    canvas.delete("pdf")
    canvas.create_image(offset_x, offset_y, anchor="nw",
                        image=rendered_img, tags="pdf")

def render_overlays():
    canvas.delete("overlay")
    for b in bubbles:
        if b["page"] == current_page_index:
            x = b["x"] * zoom + offset_x
            y = b["y"] * zoom + offset_y
            r = b["r"] * zoom
            canvas.create_oval(
                x-r, y-r, x+r, y+r,
                outline="red", width=r/10, tags="overlay"
            )
            canvas.create_text(
                x, y, text=str(b["no"]),
                font=("Arial", int(r)),
                fill="red", tags="overlay"
            )

def render(force=False):
    if not doc:
        return

    if force or rendered_zoom != zoom:
        render_pdf()

    render_overlays()
    update_bubble_list()



def open_pdf():
    global PDF_IN, doc, num_pages, current_page_index, bubbles, bubble_no
    global offset_x, offset_y

    path = filedialog.askopenfilename(
        title="Select PDF",
        filetypes=[("PDF files", "*.pdf")]
    )

    if not path:
        return

    if doc:
        doc.close()

    PDF_IN = path
    doc = fitz.open(PDF_IN)
    num_pages = len(doc)
    current_page_index = 0
    bubbles.clear()
    bubble_no = 1
    offset_x = offset_y = 0

    render()

# =====================================================
# POPUP (Requirement / Tolerance)
# =====================================================
def requirement_popup():
    popup = tk.Toplevel(root)
    popup.title("FAIR Dimension Entry")
    popup.grab_set()

    result = {"ok": False}

    tk.Label(popup, text="Characteristic").grid(row=0, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="Requirement").grid(row=1, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="- Tol").grid(row=2, column=0, padx=8, pady=5, sticky="e")
    tk.Label(popup, text="+ Tol").grid(row=3, column=0, padx=8, pady=5, sticky="e")

    char = tk.Entry(popup, width=25)
    req  = tk.Entry(popup, width=20)
    neg  = tk.Entry(popup, width=10)
    pos  = tk.Entry(popup, width=10)

    char.grid(row=0, column=1)
    req.grid(row=1, column=1)
    neg.grid(row=2, column=1, sticky="w")
    pos.grid(row=3, column=1, sticky="w")

    def submit():
        if not req.get().strip():
            messagebox.showwarning("Missing", "Requirement is mandatory")
            return
        result["ok"] = True
        result["char"] = char.get().strip()
        result["req"]  = req.get().strip()
        result["neg"]  = neg.get().strip()
        result["pos"]  = pos.get().strip()
        popup.destroy()

    tk.Button(popup, text="OK", width=10, command=submit).grid(
        row=4, column=0, columnspan=2, pady=10
    )

    popup.wait_window()
    return result

# =====================================================
# ADD BUBBLE (IMMEDIATE)
# =====================================================
def add_bubble(event):
    global bubble_no

    pdf_x = (event.x - offset_x) / zoom
    pdf_y = (event.y - offset_y) / zoom

    bubbles.append({
        "page": current_page_index,
        "no": bubble_no,
        "x": pdf_x,
        "y": pdf_y,
        "r": bubble_radius_slider.get(),
        "char": "",
        "req": "",
        "neg": "",
        "pos": ""
    })

    render()

    data = requirement_popup()

    if data.get("ok"):
        bubbles[-1]["char"] = data["char"]
        bubbles[-1]["req"]  = data["req"]
        bubbles[-1]["neg"]  = data["neg"]
        bubbles[-1]["pos"]  = data["pos"]
        bubble_no += 1
    else:
        bubbles.pop()

    render()

# =====================================================
# ZOOM / PAN
# =====================================================
zoom_job = None

def zoom_canvas(event):
    global zoom, offset_x, offset_y, zoom_job

    factor = 1.25 if event.delta > 0 else 1 / 1.25
    new_zoom = zoom * factor

    if new_zoom > 10.0 or new_zoom < 0.5 :
        return

    mx, my = event.x, event.y
    offset_x = mx - factor * (mx - offset_x)
    offset_y = my - factor * (my - offset_y)
    zoom = new_zoom

    update_preview(bubble_radius_slider.get())

    if zoom_job:
        root.after_cancel(zoom_job)

    zoom_job = root.after(120, lambda: render(force=True))


def start_pan(event):
    global pan_start
    pan_start = (event.x, event.y)

def do_pan(event):
    global offset_x, offset_y, pan_start
    if pan_start:
        dx = event.x - pan_start[0]
        dy = event.y - pan_start[1]
        offset_x += dx
        offset_y += dy
        pan_start = (event.x, event.y)

        canvas.move("pdf", dx, dy)
        canvas.move("overlay", dx, dy)


def end_pan(event):
    global pan_start
    pan_start = None

# =====================================================
# PAGE NAV / UNDO
# =====================================================
def next_page():
    global current_page_index, offset_x, offset_y
    if current_page_index < num_pages - 1:
        current_page_index += 1
        offset_x = offset_y = 0
        render()

def prev_page():
    global current_page_index, offset_x, offset_y
    if current_page_index > 0:
        current_page_index -= 1
        offset_x = offset_y = 0
        render()

def undo():
    global bubble_no
    for i in reversed(range(len(bubbles))):
        if bubbles[i]["page"] == current_page_index:
            bubbles.pop(i)
            bubble_no -= 1
            break
    render()

# =====================================================
# LIST VIEW
# =====================================================
def update_bubble_list():
    bubble_listbox.delete(0, tk.END)
    bubble_listbox.insert(tk.END, "No | Char | Req | -Tol | +Tol")
    bubble_listbox.insert(tk.END, "-" * 40)

    for b in bubbles:
        if b["page"] == current_page_index:
            bubble_listbox.insert(
                tk.END,
                f"{str(b['no']).ljust(2)} | "
                f"{str(b['char']).ljust(5)} | "
                f"{str(b['req']).ljust(5)} | "
                f"{str(b['neg']).ljust(4)} | "
                f"{str(b['pos']).ljust(4)}"
            )

# =====================================================
# SAVE BUBBLED PDF
# =====================================================
def save_pdf():
    if not bubbles:
        messagebox.showwarning("No data", "No bubbles to export")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    default_name = f"FAIR_QA_drawing_{timestamp}.pdf"

    PDF_OUT = filedialog.asksaveasfilename(
        initialfile=default_name,
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        title="Save QA Report As"
    )

    if not PDF_OUT:
        return

    out = fitz.open()
    for i in range(num_pages):
        p = out.new_page(width=doc[i].rect.width, height=doc[i].rect.height)
        p.show_pdf_page(p.rect, doc, i)
        for b in bubbles:
            if b["page"] == i:
                x, y, r = b["x"], b["y"], b["r"]
                p.draw_oval(fitz.Rect(x-r, y-r, x+r, y+r), color=(1,0,0), width=r/10)
                p.insert_text(fitz.Point(x-(r/3), y+(r/3)), str(b["no"]), fontsize=r, color=(1,0,0))
    out.save(PDF_OUT)
    out.close()
    messagebox.showinfo("Saved", f"Bubbled drawing saved as {PDF_OUT}")


# =====================================================
# SAVE REPORT (UNCHANGED)
# =====================================================
def save_report():
    if not bubbles:
        messagebox.showwarning("No data", "No bubbles to export")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    default_name = f"FAIR_Report_{timestamp}.xlsx"

    report_file = filedialog.asksaveasfilename(
        title="Save FAIR Report",
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not report_file:
        return

    shutil.copy(TEMPLATE_XLSX, report_file)
    wb = load_workbook(report_file)
    ws = wb.active

    # The template has 15 data rows (8-22). Row 23 participates in the merged
    # Legends block (A23:B24, C23:M24), so we must insert before that at row 23.
    START_ROW = 8
    TEMPLATE_ROWS = 15
    LAST_TEMPLATE_ROW = START_ROW + TEMPLATE_ROWS - 1  # 22
    INSERT_AT = LAST_TEMPLATE_ROW + 1                  # 23 (just before merged legends)
    template_row_idx = LAST_TEMPLATE_ROW
    footer_rows = {24: ws.row_dimensions[24].height,
                   25: ws.row_dimensions[25].height,
                   26: ws.row_dimensions[26].height}

    def copy_row_style(src_row_idx, dest_row_idx, clear_values=True):
        # Copy styles across all columns to preserve borders/alignment
        for col_idx in range(1, ws.max_column + 1):
            src_cell = ws.cell(row=src_row_idx, column=col_idx)
            dest_cell = ws.cell(row=dest_row_idx, column=col_idx)
            dest_cell._style = copy(src_cell._style)
            if clear_values:
                dest_cell.value = None
        # Copy row height to keep grid consistent
        ws.row_dimensions[dest_row_idx].height = ws.row_dimensions[src_row_idx].height

    bubble_count = len(bubbles)
    extra_rows = max(0, bubble_count - TEMPLATE_ROWS)

    # Insert new rows just above the merged footer block (row 23+)
    if extra_rows > 0:
        ws.insert_rows(INSERT_AT, amount=extra_rows)

        for i in range(extra_rows):
            target_row = INSERT_AT + i
            copy_row_style(template_row_idx, target_row)

        # Re-merge the footer block (Legends/Observations/Sign-off) so it
        # stays intact after the shift, and restore footer row heights.
        footer_merges = [
            "A24:B24", "C24:M24",
            "D25:M25",
            "A26:B26", "C26:D26", "E26:F26", "G26:H26", "I26:J26", "K26:M26",
        ]
        for rng in footer_merges:
            if rng in ws.merged_cells:
                ws.unmerge_cells(rng)
        for rng in footer_merges:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            ws.merge_cells(
                start_row=min_row + extra_rows,
                start_column=min_col,
                end_row=max_row + extra_rows,
                end_column=max_col,
            )
        for src_row, height in footer_rows.items():
            ws.row_dimensions[src_row + extra_rows].height = height

    # Refresh styles on all data rows (including originals) to keep borders/alignments
    style_src_row = START_ROW  # first data row has canonical formatting
    total_rows = bubble_count + 1
    for r in range(START_ROW, START_ROW + total_rows):
        copy_row_style(style_src_row, r, clear_values=False)

    # WRITE DATA (flows naturally)
    row = START_ROW
    for b in bubbles:
        ws.cell(row=row, column=1).value = b["page"] + 1
        ws.cell(row=row, column=2).value = b["no"]
        ws.cell(row=row, column=3).value = b['char']
        ws.cell(row=row, column=4).value = b["req"]
        ws.cell(row=row, column=5).value = b["neg"]
        ws.cell(row=row, column=6).value = b["pos"]
        row += 1

    wb.save(report_file)
    messagebox.showinfo("Saved", f"FAIR report created:\n{report_file}")


# =====================================================
# UI
# =====================================================
root = tk.Tk()
root.title("FAIR-y")

toolbar = tk.Frame(root)
toolbar.pack(fill="x")

tk.Button(toolbar, text="Open PDF", command=open_pdf).pack(side="left")
tk.Button(toolbar, text="Prev Page", command=prev_page).pack(side="left")
tk.Button(toolbar, text="Next Page", command=next_page).pack(side="left")
tk.Button(toolbar, text="Undo Bubble", command=undo).pack(side="left")
tk.Button(toolbar, text="Save PDF", command=save_pdf).pack(side="left")
tk.Button(toolbar, text="Save Report", command=save_report).pack(side="left")

bubble_radius_slider = tk.Scale(toolbar, from_=3, to=25, orient="horizontal", label="Bubble Size")
bubble_radius_slider.set(6)
bubble_radius_slider.pack(side="right")

preview_canvas = tk.Canvas(toolbar, width=50, height=50, bg="gray")
preview_canvas.pack(side="right", padx=5)

def update_preview(val):
    preview_canvas.delete("all")
    r = int(val) * zoom
    preview_canvas.create_oval(25-r, 25-r, 25+r, 25+r, outline="red", width=2)

bubble_radius_slider.config(command=update_preview)
update_preview(bubble_radius_slider.get())

bubble_listbox = tk.Listbox(root, height=7, font=("Consolas", 10))
bubble_listbox.pack(fill="x", padx=5, pady=3)

canvas = tk.Canvas(root, bg="gray")
canvas.pack(fill="both", expand=True)

canvas.bind("<Button-3>", add_bubble)
canvas.bind("<Button-1>", start_pan)
canvas.bind("<B1-Motion>", do_pan)
canvas.bind("<ButtonRelease-1>", end_pan)
canvas.bind("<MouseWheel>", zoom_canvas)

render()
root.mainloop()
doc.close()
